"""Export d'une analyse de CDC vers Excel (.xlsx) ou Markdown (.md)."""
from __future__ import annotations

import io
from datetime import datetime
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


STATUS_LABEL = {
    "covered": "Couvert",
    "partial": "Partiel",
    "missing": "Manquant",
    "ambiguous": "Ambigu",
}

STATUS_COLOR = {
    "covered": "C8F1D6",
    "partial": "FCE7B1",
    "missing": "F6C7C7",
    "ambiguous": "E2E8F0",
}


def _flatten_evidence(evidence: List[Any]) -> str:
    if not evidence:
        return ""
    parts: List[str] = []
    for e in evidence:
        if isinstance(e, str):
            parts.append(e)
        elif isinstance(e, dict):
            parts.append(str(e.get("text") or e.get("content") or ""))
    return "\n— ".join(["", *parts]).lstrip("\n").strip()


def _flatten_sources(sources: List[Any]) -> str:
    if not sources:
        return ""
    parts: List[str] = []
    for s in sources:
        if not isinstance(s, dict):
            continue
        src = s.get("source") or "?"
        page = s.get("page")
        if page is not None:
            parts.append(f"{src} (p.{page})")
        else:
            parts.append(str(src))
    return "; ".join(parts)


def build_xlsx(filename: str, report: Dict[str, Any]) -> bytes:
    """Produit un .xlsx avec deux feuilles : Synthèse + Exigences."""
    wb = Workbook()

    # --- Feuille 1 : Synthèse ---
    ws1 = wb.active
    ws1.title = "Synthèse"

    title_font = Font(name="Calibri", size=14, bold=True, color="0B1B2B")
    label_font = Font(name="Calibri", size=11, bold=True)
    ws1["A1"] = "Analyse de cahier des charges"
    ws1["A1"].font = title_font
    ws1["A2"] = filename
    ws1["A2"].font = Font(italic=True, color="6B7A90")
    ws1["A3"] = f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws1["A3"].font = Font(italic=True, color="6B7A90")

    summary = report.get("summary") or {}
    rows = [
        ("Total exigences", summary.get("total", 0)),
        ("Couvertes", summary.get("covered", 0)),
        ("Partielles", summary.get("partial", 0)),
        ("Manquantes", summary.get("missing", 0)),
        ("Ambiguës", summary.get("ambiguous", 0)),
        ("Couverture", f"{summary.get('coverage_percent', 0):.1f}%"),
    ]
    for i, (label, value) in enumerate(rows, start=5):
        ws1.cell(row=i, column=1, value=label).font = label_font
        ws1.cell(row=i, column=2, value=value)

    ws1.column_dimensions["A"].width = 24
    ws1.column_dimensions["B"].width = 18

    # --- Feuille 2 : Exigences ---
    ws2 = wb.create_sheet("Exigences")

    # Bandeau d'avertissement en haut (lignes 1 à 3) : explique aux humains
    # comment remplir les colonnes de correction et NE PAS toucher la colonne ID.
    notice_fill = PatternFill("solid", fgColor="FFF4D6")  # jaune doux
    notice_font = Font(bold=True, color="6B4F00", size=11)
    notice_lines = [
        "⚠ Pour CORRIGER ces verdicts depuis Excel : remplissez les 3 dernières "
        "colonnes (Verdict humain · Description corrigée · Notes internes), "
        "puis ré-importez ce fichier dans le rapport via le bouton "
        "« Importer corrections ».",
        "Verdict humain : choisir parmi covered / partial / missing (case "
        "vide = aucune correction). Description corrigée : texte libre.",
        "⚠ NE PAS modifier la colonne « ID » (1ʳᵉ colonne). Elle sert au "
        "ré-import : un ID modifié ou inconnu fera ignorer la ligne.",
    ]
    NOTICE_ROWS = len(notice_lines)
    LAST_COL = 14  # 11 cols actuelles + 3 cols correction
    for i, msg in enumerate(notice_lines, start=1):
        ws2.merge_cells(
            start_row=i, start_column=1, end_row=i, end_column=LAST_COL
        )
        cell = ws2.cell(row=i, column=1, value=msg)
        cell.fill = notice_fill
        cell.font = notice_font
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
        ws2.row_dimensions[i].height = 32

    HEADER_ROW = NOTICE_ROWS + 1  # ligne 4 par défaut

    headers = [
        "ID",
        "Titre",
        "Catégorie",
        "Priorité",
        "Statut",
        "Verdict",
        "Description",
        "Critères d'acceptation",
        "Localisation source",
        "Preuves",
        "Sources retrouvées",
        "🖋 Verdict humain",
        "🖋 Description corrigée",
        "🖋 Notes internes",
    ]
    header_fill = PatternFill("solid", fgColor="2E6BE6")
    correction_header_fill = PatternFill("solid", fgColor="7C3AED")  # violet
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, h in enumerate(headers, start=1):
        cell = ws2.cell(row=HEADER_ROW, column=col_idx, value=h)
        cell.fill = (
            correction_header_fill if col_idx >= 12 else header_fill
        )
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", horizontal="left")
    ws2.row_dimensions[HEADER_ROW].height = 22

    # Cellules ID (col A) en gris doux pour signaler "ne pas modifier".
    id_fill = PatternFill("solid", fgColor="EEF1F5")
    id_font = Font(name="Consolas", size=10, color="6B7A90")
    correction_fill = PatternFill("solid", fgColor="F5F0FF")  # violet très clair

    requirements = report.get("requirements") or []
    DATA_START = HEADER_ROW + 1
    for offset, r in enumerate(requirements):
        row = DATA_START + offset
        status = (r.get("status") or "").lower()
        # ID en gris (visuellement "verrouillé")
        id_cell = ws2.cell(row=row, column=1, value=r.get("id", ""))
        id_cell.fill = id_fill
        id_cell.font = id_font
        ws2.cell(row=row, column=2, value=r.get("title", ""))
        ws2.cell(row=row, column=3, value=r.get("category", ""))
        ws2.cell(row=row, column=4, value=r.get("priority", ""))
        status_cell = ws2.cell(
            row=row, column=5, value=STATUS_LABEL.get(status, status)
        )
        if status in STATUS_COLOR:
            status_cell.fill = PatternFill("solid", fgColor=STATUS_COLOR[status])
        ws2.cell(row=row, column=6, value=r.get("verdict", ""))
        ws2.cell(row=row, column=7, value=r.get("description", ""))
        ac = r.get("acceptance_criteria") or []
        ws2.cell(row=row, column=8, value="\n".join(f"• {x}" for x in ac))
        ws2.cell(row=row, column=9, value=r.get("source_location", ""))
        ws2.cell(row=row, column=10, value=_flatten_evidence(r.get("evidence") or []))
        ws2.cell(row=row, column=11, value=_flatten_sources(r.get("sources") or []))
        # 3 colonnes vides pour la correction humaine, fond violet très clair
        for col in (12, 13, 14):
            c = ws2.cell(row=row, column=col, value="")
            c.fill = correction_fill

    # Data validation : dropdown sur la colonne "Verdict humain" (col 12).
    if requirements:
        dv = DataValidation(
            type="list",
            formula1='"covered,partial,missing"',
            allow_blank=True,
            showDropDown=False,  # affiche bien le dropdown
        )
        dv.error = "Valeur invalide : utilisez covered, partial, missing ou laissez vide."
        dv.errorTitle = "Verdict humain"
        dv.prompt = "Sélectionnez covered, partial ou missing — ou laissez vide."
        dv.promptTitle = "Verdict humain"
        ws2.add_data_validation(dv)
        last_data_row = DATA_START + len(requirements) - 1
        dv.add(f"L{DATA_START}:L{last_data_row}")

    widths = [10, 36, 22, 12, 14, 50, 50, 40, 22, 50, 40, 16, 50, 30]
    for idx, w in enumerate(widths, start=1):
        ws2.column_dimensions[get_column_letter(idx)].width = w
    # Freeze panes : on fige les lignes du bandeau + l'en-tête, et la 1ʳᵉ
    # colonne (ID) pour qu'elle reste visible quand on scrolle horizontalement.
    ws2.freeze_panes = ws2.cell(row=HEADER_ROW + 1, column=2).coordinate
    for row in ws2.iter_rows(min_row=DATA_START):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Import corrections depuis le même format Excel (3 dernières colonnes)
# ---------------------------------------------------------------------------


# Index 1-based des colonnes attendues dans le fichier Excel exporté ci-dessus
CORRECTION_COL_VERDICT = 12
CORRECTION_COL_ANSWER = 13
CORRECTION_COL_NOTES = 14
EXPORT_LAST_COL = 14
VALID_HUMAN_VERDICTS = {"covered", "partial", "missing"}


def parse_corrections_xlsx(data: bytes) -> Dict[str, Any]:
    """Parse un .xlsx exporté par build_xlsx et extrait les corrections
    humaines (lignes avec Verdict humain non vide).

    Renvoie un dict :
      {
        "rows": [{requirement_id, verdict, answer, notes}, ...],
        "skipped": [{row_excel, reason, raw_id}, ...],
      }

    Le caller (endpoint d'import) se charge ensuite de valider que les IDs
    existent bien dans l'analyse cible et d'appeler workspace.upsert_correction.
    """
    from openpyxl import load_workbook  # late import — runtime only

    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    if "Exigences" not in wb.sheetnames:
        raise ValueError(
            "Fichier non reconnu : feuille « Exigences » introuvable. "
            "Avez-vous bien exporté ce fichier depuis le rapport CDC ?"
        )
    ws = wb["Exigences"]

    rows: List[Dict[str, Any]] = []
    skipped: List[Dict[str, Any]] = []
    # On scanne toutes les lignes : on ignore celles dont la 1ʳᵉ cellule est
    # vide (notice fusionnée + en-tête) ou commence par un emoji/texte de
    # bandeau. La détection des données utiles se fait via "ID non vide
    # ET col verdict humain non vide".
    for excel_row in ws.iter_rows(min_row=1, values_only=True):
        if not excel_row:
            continue
        if len(excel_row) < EXPORT_LAST_COL:
            continue
        rid_raw = excel_row[0]
        verdict_raw = excel_row[CORRECTION_COL_VERDICT - 1]
        if rid_raw is None or str(rid_raw).strip() == "":
            continue  # ligne du bandeau, en-tête, ou ligne vide
        rid = str(rid_raw).strip()
        # En-tête : ID = "ID" littéralement
        if rid.lower() == "id":
            continue
        verdict = (
            str(verdict_raw).strip().lower() if verdict_raw is not None else ""
        )
        if not verdict:
            continue  # pas de correction sur cette ligne
        if verdict not in VALID_HUMAN_VERDICTS:
            skipped.append({
                "raw_id": rid,
                "reason": f"Verdict invalide : '{verdict}'.",
            })
            continue
        answer_raw = excel_row[CORRECTION_COL_ANSWER - 1]
        notes_raw = excel_row[CORRECTION_COL_NOTES - 1]
        answer = str(answer_raw).strip() if answer_raw else ""
        notes = str(notes_raw).strip() if notes_raw else ""
        if not answer:
            skipped.append({
                "raw_id": rid,
                "reason": "Description corrigée vide (obligatoire).",
            })
            continue
        rows.append({
            "requirement_id": rid,
            "verdict": verdict,
            "answer": answer,
            "notes": notes or None,
        })
    wb.close()
    return {"rows": rows, "skipped": skipped}


def build_markdown(filename: str, report: Dict[str, Any]) -> str:
    """Produit un rapport Markdown synthétique de l'analyse."""
    summary = report.get("summary") or {}
    requirements = report.get("requirements") or []
    lines: List[str] = []
    lines.append(f"# Analyse de cahier des charges — {filename}")
    lines.append("")
    lines.append(f"_Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}_")
    lines.append("")

    lines.append("## Synthèse")
    lines.append("")
    lines.append("| Indicateur | Valeur |")
    lines.append("|---|---|")
    lines.append(f"| Total exigences | {summary.get('total', 0)} |")
    lines.append(f"| Couvertes | {summary.get('covered', 0)} |")
    lines.append(f"| Partielles | {summary.get('partial', 0)} |")
    lines.append(f"| Manquantes | {summary.get('missing', 0)} |")
    lines.append(f"| Ambiguës | {summary.get('ambiguous', 0)} |")
    lines.append(f"| **Couverture** | **{summary.get('coverage_percent', 0):.1f}%** |")
    lines.append("")

    # Group by status for readability
    groups: Dict[str, List[Dict[str, Any]]] = {
        "covered": [],
        "partial": [],
        "missing": [],
        "ambiguous": [],
    }
    for r in requirements:
        st = (r.get("status") or "").lower()
        groups.setdefault(st, []).append(r)

    headers = [
        ("covered", "Exigences couvertes"),
        ("partial", "Exigences partielles"),
        ("missing", "Exigences manquantes"),
        ("ambiguous", "Exigences ambiguës"),
    ]
    for key, h in headers:
        items = groups.get(key) or []
        if not items:
            continue
        lines.append(f"## {h} ({len(items)})")
        lines.append("")
        for r in items:
            lines.append(f"### {r.get('id', '')} — {r.get('title', '')}")
            if r.get("category"):
                lines.append(f"**Catégorie** : {r['category']}  ")
            if r.get("priority"):
                lines.append(f"**Priorité** : {r['priority']}  ")
            if r.get("source_location"):
                lines.append(f"**Localisation source** : {r['source_location']}  ")
            lines.append("")
            if r.get("description"):
                lines.append(r["description"])
                lines.append("")
            if r.get("verdict"):
                lines.append(f"> **Verdict** — {r['verdict']}")
                lines.append("")
            ac = r.get("acceptance_criteria") or []
            if ac:
                lines.append("**Critères d'acceptation** :")
                for x in ac:
                    lines.append(f"- {x}")
                lines.append("")
            ev = r.get("evidence") or []
            if ev:
                lines.append("**Preuves** :")
                for e in ev:
                    if isinstance(e, str):
                        lines.append(f"- {e}")
                    elif isinstance(e, dict):
                        txt = e.get("text") or e.get("content") or ""
                        lines.append(f"- {txt}")
                lines.append("")
            srcs = r.get("sources") or []
            if srcs:
                lines.append("**Sources retrouvées** :")
                for s in srcs:
                    if not isinstance(s, dict):
                        continue
                    src = s.get("source") or "?"
                    page = s.get("page")
                    score = s.get("score")
                    extra = []
                    if page is not None:
                        extra.append(f"p.{page}")
                    if isinstance(score, (int, float)):
                        extra.append(f"score {score:.3f}")
                    suffix = f" ({', '.join(extra)})" if extra else ""
                    lines.append(f"- {src}{suffix}")
                lines.append("")
            lines.append("---")
            lines.append("")

    return "\n".join(lines)
